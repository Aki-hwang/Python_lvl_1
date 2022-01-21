import urllib.request as req
from bs4 import BeautifulSoup
import openpyxl #엑셀에 넣기 위해
import os
from openpyxl.drawing.image import Image #앨범을 넣기 위해 부러와
#엑셀파일 존재하는지 확인 위해 OS 블러옴

if not os.path.exists("./멜론음원차트_오늘.xlsx"):
    book = openpyxl.Workbook()
    book.save("./멜론음원차트_오늘.xlsx")

headers = req.Request("https://www.melon.com/chart/index.htm", headers={"user-Agent": "Mozilla/5.0"})
code = req.urlopen(headers)
soup = BeautifulSoup(code, "lxml")
title = soup.select("div.ellipsis.rank01 a") #속성값에 띄어쓰기 있으면 점으로
name = soup.select("div.ellipsis.rank02 > span.checkEllipsis")
album = soup.select("div.ellipsis.rank03 a")
image = soup.select("a.image_typeAll img")

book = openpyxl.load_workbook("./멜론음원차트_오늘.xlsx") #먼저 불러와야해
# sheet = book["Sheet1"]
sheet = book.active #자동으로 열리는 그 시트를 사용하겠다
sheet.column_dimensions["A"].width = 15
sheet.column_dimensions["B"].width = 50
sheet.column_dimensions["C"].width = 30
sheet.column_dimensions["D"].width = 35

row_num = 1
#이미지 저정할 폴더 만들기
if not os.path.exists("./멜론이미지"):
    os.mkdir("./멜론이미지")

for i in range(len(title)):
    req.urlretrieve(image[i].attrs["src"], "./멜론이미지/{}.png".format(row_num))
    print(title[i].string, name[i].text, album[i].string, image[i].attrs["src"])
    img_for_ex = Image("./멜론이미지/{}.png".format(row_num)) #이미지 함수가 엑셀에 넣을 수 있게 변환을 해죠
    sheet.row_dimensions[row_num].height = 95
    sheet.add_image(img_for_ex,"A{}".format(row_num)) #이미지 넣는 방법
    sheet.cell(row=row_num, column=2).value = title[i].string
    sheet.cell(row=row_num, column=3).value = name[i].text
    sheet.cell(row=row_num, column=4).value = album[i].string
    book.save("./멜론음원차트_오늘.xlsx")
    row_num += 1



