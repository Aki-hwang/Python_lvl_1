import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Color, Side
from openpyxl.chart import Reference, Series, LineChart, BarChart
import pandas as pd

df = pd.read_excel("./아파트분양가격.xlsx")
locations = df["지역명"].unique()#중복값 지워주는 명령

for loc in locations:
    write = pd.ExcelWriter("./엑셀/아파트분양가격_{}.xlsx".format(loc))  # 시트 네임을 넣기 위해 write 사용
    df_location = df[df["지역명"] == loc]
    years = df_location["연도"].unique()
    for y in years:
        df_year = df_location[df_location["연도"] == y] #특정 지역명 중 특정 연도 추출
        df_result = df_year.sort_values(by="월", ascending=True)
        df_result.to_excel(write, sheet_name="{}년".format(y), index=None)
    write.save() #하나의 엑셀파일이 완성
    print("{}지역 엑셀파일 분리 완료".format(loc))

    #판다스로 (데이터 정렬등) 원하는 업무를 다하고 서식변경등을 하려면 다시 파일 불러오는게 나아

#엑셀 셀서식 만들기
user_font = Font(name="맑은 고딕", size=12, bold=True)
user_alignment = Alignment(horizontal="center")
orange_color = PatternFill(patternType="solid", fgColor=Color("FE9A2E"))
gray_color = PatternFill(patternType="solid", fgColor=Color("BDBDBD"))
user_boder = Border(left=Side(style="thin") , right=Side(style="thin") , top=Side(style="thin") , bottom=Side(style="thin"))


for loc in locations:
    book = openpyxl.load_workbook("./엑셀/아파트분양가격_{}.xlsx".format(loc))
    for sheetname in book.sheetnames:    #시트네임을 리스트 자료형으로 반환해죠 --> 포문에 사용
        sheet = book[sheetname]
        #데이터에 헤더에 셀서식 지정
        for row in sheet["A1:E1"]:
            for cell in row:
                cell.font = user_font
                cell.alignment = user_alignment
                cell.fill = orange_color
                cell.border = user_boder
        #데이터에 내용에 셀서식 지정
        for row in sheet["A2:E{}".format(sheet.max_row)]: #sheet.max_low 최대 행번호를 반환해
            for cell in row:
                cell.alignment = user_alignment
                cell.fill = orange_color
                cell.border = user_boder
        chart = BarChart()
        chart.title = "{}지역 {}년도 아파트 분양가격".format(loc, sheetname)
        value = Reference(sheet, range_string="{}!E2:E{}".format(sheetname,sheet.max_row))
        value_series = Series(value,title="분양가격")
        chart.append(value_series)
        sheet.add_chart(chart,"G1")
    book.save("./엑셀/아파트분양가격_{}.xlsx".format(loc))
    print("{}지역 엑셀 서식 지정 완료.xlsx".format(loc))