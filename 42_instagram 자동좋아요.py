from selenium import webdriver
import time
import random
import openpyxl
import os
import unicodedata #모음 자음 분리될때

if not os.path.exists("./인스타그램.xlsx"):
    book = openpyxl.Workbook()
    book.save("./인스타그램.xlsx")
book = openpyxl.load_workbook("./인스타그램.xlsx") #엑셀 파일 불러오는 코드
sheet = book.active

hash_tag = input("해시태그 입력 >>>")
browser = webdriver.Chrome()
browser.get("https://www.instagram.com/accounts/login/")
time.sleep(3)

id = browser.find_element_by_name("username") #네임속성값을 가져와
id.send_keys("aki_aki_sue")
pwd = browser.find_element_by_name("password")
pwd.send_keys("Hwang1234!@")
browser.find_element_by_css_selector("div.qF0y9.Igw0E.IwRSH.eGOV_._4EzTm.bkEs3.CovQj.jKUp7.DhRcB").click()
time.sleep(4)

url = "https://www.instagram.com/explore/tags/{}/".format(hash_tag)
browser.get(url)
time.sleep(5)
#자동 좋아요 동작 시키기
# 첫번째 사진 클릭하기
browser.find_element_by_css_selector("div._9AhH0").click()
time.sleep(3)
row_num = 0
while True:
    #좋아요 요소
    like = browser.find_element_by_css_selector("section.ltpMr.Slqrh button.wpO6b svg._8-yf5")
    next = browser.find_element_by_css_selector("button.wpO6b")
    value = like.get_attribute("aria-label")
    nick_name = browser.find_element_by_css_selector("") # 글작성자
    content = browser.find_element_by_css_selector("") # 글내용 content눈 html이라 text를 해줘야 내용이 들어가
    content.noramlize = unicodedata.normalize("NFC", content.text) #한글의 자음 모음 현상을 해결해준다
    sheet.cell(row=row_num , column=1 ).value = nick_name.text
    sheet.cell(row=row_num, column=2).value = content.noramlize.text
    row_num += 1
    book.save("./인스타그램.xlsx")

    if value == "좋아요":
        like.click()
        time.sleep(random.randint(2,5) + random.random()) #완전한 랜덤 숫자
        next.click()
        time.sleep(random.randint(2,5) + random.random()) #완전한 랜덤 숫자
    elif value == "좋아요 취소":
        next.click()
        time.sleep(random.randint(2,5) + random.random()) #완전한 랜덤 숫자
